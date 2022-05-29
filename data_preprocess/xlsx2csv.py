import os
import glob
import openpyxl
import logging
logging.basicConfig(
    level=logging.INFO,
    format= '%(asctime)s - %(levelname)s - %(message)s'
)

xlsx_pathL = glob.glob("data/Ori/*.xlsx")

for xlsx_path in xlsx_pathL:
    # xlsx_path = xlsx_pathL[0]

    dir_name = os.path.splitext(os.path.basename(xlsx_path))[0]
    dir_path = f"data/{dir_name}"

    if not os.path.exists(dir_path):
        os.mkdir(dir_path)

    logging.info(dir_name)
    # 打开xlsx
    wb = openpyxl.load_workbook(xlsx_path)
    for sheetnames in wb.sheetnames:
        sheet = wb[sheetnames]
        # line_text = ""

        data_textL = [
            ",".join([str(c.value) for c in r])+"\n"
            for r in sheet.rows
        ]
        with open(f"{dir_path}/{dir_name}_{sheetnames}.csv","w",encoding="utf-8") as fp:
            fp.writelines(data_textL)
        logging.info(sheetnames)



